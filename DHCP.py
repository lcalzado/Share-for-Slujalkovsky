#importando modulos
import pandas as pd
from openpyxl import Workbook
from netmiko import ConnectHandler
import ipaddress
import re
import pymysql

#Archivando la data.
path = "/home/this/is/a/path/to/a/directory/lease.xlsx"
path2 = "/home/this/is/a/path/to/a/directory/total2.xlsx"

#Definiendo un parametro para el prompt de la CLI.
prompt_pattern_1 = r".*.*"

#Definiendo listas en blanco para recopilar la data.
list1 = []
list2 = []
list3 = []
list4 = []
list5 = []

#Credendiales para la conexion ssh
device = {
    'device_type': 'fortinet',  #Tipo de dispositivo
    'ip': 'host',
    'username': 'user',
    'password': '************',
    }

#Credendiales para la Base de datos   
db_config = {
    'host': 'host',
    'user': 'user',
    'password': '***********',
    'database': 'data_base_name',
}

#lectura del primer archivo
data = pd.read_excel(path)

#----------------------------------------Loops--------------------------------

#-----------------------------------------------------------------------------

for index, row in data.iterrows():
    interfaz = row['Interfaz']
    list1.append(interfaz)


with ConnectHandler(**device) as ssh_client:

    for index, row in data.iterrows():
        command = row['Command']
        ssh_client.send_command('config vdom', expect_string=prompt_pattern_1)
        ssh_client.send_command('edit DHCP-1', expect_string=prompt_pattern_1)
        command_output = ssh_client.send_command(command)
        ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
        ips = re.findall(ip_pattern, command_output)
        cantidad_ips = len(ips)
        list2.append(cantidad_ips)
        
#------------------------------------------------------------------------------

#Lectura del segundo archivo.
data2 = pd.read_excel(path2)


for index, row in data2.iterrows():
    interfaz_ref = row['Interfaz_ref']
    list3.append(interfaz_ref)   


with ConnectHandler(**device) as ssh_client2:

    ssh_client2.send_command('config vdom', expect_string=prompt_pattern_1)
    ssh_client2.send_command('edit DHCP-1', expect_string=prompt_pattern_1)
    ssh_client2.send_command('config system dhcp server', expect_string=prompt_pattern_1)
    
    for index, row in data2.iterrows():
        command2 = row['Command2']
        ssh_client2.send_command(command2, expect_string=prompt_pattern_1)
        command_output2 = ssh_client2.send_command('show')
        ip_pattern2 = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'
        ips2 = re.findall(ip_pattern2, command_output2)
        start_ip = ips2[2]
        end_ip = ips2[3]
        ip1 = ipaddress.IPv4Address(start_ip)
        ip2 = ipaddress.IPv4Address(end_ip)
        cantidad_ips2 = int(ip2) - int(ip1) + 1
        list4.append(cantidad_ips2)
        ssh_client2.send_command('next', expect_string=prompt_pattern_1)
        
#------------------------------------------------------------------------------       

#Creando la tabla dinamica.

pvot = {'ColumnaA': list3, 'ColumnaB': list4}
df = pd.DataFrame(pvot)
# Crear una tabla dinámica
tabla_dinamica = df.pivot_table(values='ColumnaB', index='ColumnaA', aggfunc='sum')
orden_deseado = list1
# Reindexar la tabla dinámica para mantener el orden
tabla_dinamica = tabla_dinamica.reindex(orden_deseado)
# Convertir los valores de la tabla dinámica en una nueva lista
new_list = tabla_dinamica['ColumnaB'].tolist()

#------------------------------------------------------------------------------

#Obteniendo el %
for valor1, valor2 in zip(list2, new_list):
    resultado = (valor1 / valor2) * 100
    resultado_redondeado = round(resultado, 2)
    list5.append(resultado_redondeado)

#------------------------------------------------------------------------------

# Inicializar un objeto Workbook para guardar los resultados
result_workbook = Workbook()
result_sheet = result_workbook.active

# Establecer los títulos de las columnas
result_sheet['A1'] = 'Interfaz'
result_sheet['B1'] = 'Lease'
result_sheet['C1'] = 'Total_allowed_ips'
result_sheet['D1'] = 'Utilization(%)'

for i in range(len(list1)):
    result_sheet.cell(row=i + 2, column=1, value=list1[i])

for j in range(len(list2)):
    result_sheet.cell(row=j + 2, column=2, value=list2[j])

for k in range(len(new_list)):
    result_sheet.cell(row=k + 2, column=3, value=new_list[k])

for l in range(len(list5)):
    result_sheet.cell(row=l + 2, column=4, value=list5[l])

result_workbook.save("/home/this/is/a/path/to/a/directory/resultados.xlsx")
result_workbook.close()

connection = pymysql.connect(**db_config)

try:
    with connection.cursor() as cursor:
        for i in range(len(list1)):
            # Insertar o actualizar registros en la tabla dhcp_report
            insert_query = '''
            INSERT INTO dhcp_report (Interfaz_name, Lease_count, Total_IPs, Utilization)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
            Interfaz_name = VALUES(Interfaz_name), Lease_count = VALUES(Lease_count), Total_IPs = VALUES(Total_IPs), Utilization = VALUES(Utilization)
            '''
            cursor.execute(insert_query, (list1[i], list2[i], new_list[i], list5[i]))
        
        connection.commit()

finally:
    connection.close()